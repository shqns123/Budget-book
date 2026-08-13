from pathlib import Path
import sys
from openpyxl import load_workbook

sys.stdout.reconfigure(encoding="utf-8")
source = Path(r"C:\Users\Woojin\Downloads\가계부 .xlsx")
workbook = load_workbook(source, read_only=True, data_only=True)

print("SHEETS")
for sheet in workbook.worksheets:
    print(f"{sheet.title}\t{sheet.max_row}\t{sheet.max_column}")

for name in ("가계부 기록 ✍️", "가계부 설정 ✍️"):
    sheet = workbook[name]
    print(f"\n[{name}]")
    count = 0
    for row in sheet.iter_rows(values_only=True):
        if any(value is not None and str(value).strip() for value in row):
            print(repr(row[:12]))
            count += 1
            if count >= 35:
                break

settings = workbook["가계부 설정 ✍️"]
print("\n[가계부 설정 전체 비어있지 않은 행]")
for index, row in enumerate(settings.iter_rows(values_only=True), start=1):
    values = tuple(row[:16])
    if any(value is not None and str(value).strip() for value in values):
        print(index, repr(values))

print("\n[계좌/자산 입력 영역]")
for index, row in enumerate(settings.iter_rows(min_row=12, max_row=80, min_col=14, max_col=30, values_only=True), start=12):
    if any(value is not None and str(value).strip() for value in row):
        print(index, repr(row))
