from collections import Counter
from datetime import date, datetime
from pathlib import Path
import json
import sys

from openpyxl import load_workbook

sys.stdout.reconfigure(encoding="utf-8")

ROOT = Path(__file__).resolve().parents[1]
source = Path(r"C:\Users\Woojin\Downloads\가계부 .xlsx")
payload = json.loads((ROOT / "data" / "imported-ledger.json").read_text(encoding="utf-8"))
workbook = load_workbook(source, read_only=True, data_only=True)
sheet = workbook["가계부 기록 ✍️"]

source_rows = []
for row in sheet.iter_rows(values_only=True):
    transaction_date, kind, memo, amount = row[1], str(row[2] or "").strip(), str(row[5] or "").strip(), row[6]
    if isinstance(transaction_date, (datetime, date)) and kind in {"수입", "지출", "이동"} and memo and amount is not None:
        deposit, withdrawal = str(row[7] or "").strip(), str(row[8] or "").strip()
        primary, destination = (deposit, "") if kind == "수입" else (withdrawal, "") if kind == "지출" else (withdrawal, deposit)
        source_rows.append((transaction_date.date().isoformat(), kind, memo, int(round(float(amount))), primary or "미지정 계좌", destination or None))

accounts_by_id = {account["id"]: account["name"] for account in payload["accounts"]}
imported_rows = [(item["date"], {"income": "수입", "expense": "지출", "transfer": "이동"}[item["type"]], item["memo"], item["amount"], accounts_by_id[item["accountId"]], accounts_by_id[item["toAccountId"]] if item["toAccountId"] else None) for item in payload["transactions"]]

assert Counter(source_rows) == Counter(imported_rows), "원본과 이관 거래 내역이 다릅니다."
assert len(source_rows) == payload["summary"]["transactionCount"]
assert len(payload["accounts"]) == payload["summary"]["accountCount"]
assert len(payload["categories"]) == payload["summary"]["categoryCount"]

print(json.dumps({
    "result": "verified",
    "transactions": len(source_rows),
    "accounts": len(payload["accounts"]),
    "categories": len(payload["categories"]),
    "period": [min(item[0] for item in source_rows), max(item[0] for item in source_rows)],
}, ensure_ascii=False, indent=2))
