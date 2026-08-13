"""Build a deterministic application seed from the user-authorized source workbook."""

from collections import OrderedDict
from datetime import date, datetime
from pathlib import Path
import json
import re
import sys
import uuid

from openpyxl import load_workbook

sys.stdout.reconfigure(encoding="utf-8")

ROOT = Path(__file__).resolve().parents[1]
SOURCE = Path(r"C:\Users\Woojin\Downloads\가계부 .xlsx")
OUTPUT = ROOT / "data" / "imported-ledger.json"
NAMESPACE = uuid.UUID("4e8d535c-650c-48cc-893c-0f93bc099921")


def clean(value):
    if value is None:
        return ""
    return str(value).strip()


def stable_id(kind, value):
    return str(uuid.uuid5(NAMESPACE, f"{kind}:{value}"))


def account_type(classification, minor):
    if classification == "자산":
        if minor == "현금":
            return "cash"
        if minor in {"예적금", "주택청약", "주식", "외화예금", "퇴직연금"}:
            return "savings"
        return "bank"
    if minor.startswith("신용카드"):
        return "card"
    return "loan"


def account_classification(classification, major):
    if classification == "자산":
        return "asset"
    return "short_liability" if major == "🏦단기부채" else "long_liability"


workbook = load_workbook(SOURCE, read_only=True, data_only=True)
settings = workbook["가계부 설정 ✍️"]
records = workbook["가계부 기록 ✍️"]

settings_rows = list(settings.iter_rows(values_only=True))
start_year = int(settings_rows[7][2])
fiscal_start_month = int(re.search(r"\d+", clean(settings_rows[8][2])).group())

accounts = []
for row in settings_rows[12:]:
    code = row[14] if len(row) > 14 else None
    classification = clean(row[15] if len(row) > 15 else None)
    major = clean(row[16] if len(row) > 16 else None)
    minor = clean(row[17] if len(row) > 17 else None)
    name = clean(row[18] if len(row) > 18 else None)
    balance = row[19] if len(row) > 19 else None
    memo = clean(row[20] if len(row) > 20 else None)
    hidden = bool(row[21]) if len(row) > 21 and row[21] is not None else False
    if not code or not classification or not name:
        continue
    code = str(int(code)) if isinstance(code, float) and code.is_integer() else clean(code)
    accounts.append({
        "id": stable_id("account", code),
        "accountCode": code,
        "name": name,
        "type": account_type(classification, minor),
        "classification": account_classification(classification, major),
        "majorCategory": major,
        "minorCategory": minor,
        "openingBalance": int(balance or 0),
        "memo": memo or None,
        "isHidden": hidden,
        "paymentDay": None,
        "color": "#17191c",
    })

account_by_name = {account["name"]: account for account in accounts}
missing_account_names = set()

categories = OrderedDict()
current_type = ""
for row in settings_rows[32:80]:
    kind = clean(row[1] if len(row) > 1 else None)
    if kind in {"수입", "지출"}:
        current_type = "income" if kind == "수입" else "expense"
    major = clean(row[2] if len(row) > 2 else None)
    if not current_type or not major:
        continue
    for minor_index, fixed_index in ((4, 5), (6, 7), (8, 9), (10, 11)):
        minor = clean(row[minor_index] if len(row) > minor_index else None)
        if not minor:
            continue
        fixed = bool(row[fixed_index]) if len(row) > fixed_index else False
        categories[(current_type, major, minor)] = fixed

transactions = []
for source_row, row in enumerate(records.iter_rows(values_only=True), start=1):
    transaction_date = row[1] if len(row) > 1 else None
    kind = clean(row[2] if len(row) > 2 else None)
    if not isinstance(transaction_date, (datetime, date)) or kind not in {"수입", "지출", "이동"}:
        continue
    major = clean(row[3] if len(row) > 3 else None)
    minor = clean(row[4] if len(row) > 4 else None)
    memo = clean(row[5] if len(row) > 5 else None)
    amount = row[6] if len(row) > 6 else None
    deposit_account = clean(row[7] if len(row) > 7 else None)
    withdrawal_account = clean(row[8] if len(row) > 8 else None)
    if not memo or amount is None:
        continue

    if kind == "수입":
        transaction_type = "income"
        primary_name, secondary_name = deposit_account, ""
    elif kind == "지출":
        transaction_type = "expense"
        primary_name, secondary_name = withdrawal_account, ""
    else:
        transaction_type = "transfer"
        primary_name, secondary_name = withdrawal_account, deposit_account
        major = "이동"
        minor = ""

    if not primary_name:
        primary_name = "미지정 계좌"
        missing_account_names.add(primary_name)
    if secondary_name and secondary_name not in account_by_name:
        missing_account_names.add(secondary_name)
    if primary_name not in account_by_name:
        missing_account_names.add(primary_name)

    category_key = (transaction_type, major, minor)
    if transaction_type in {"income", "expense"} and major and minor and category_key not in categories:
        categories[category_key] = False
    transaction_id = stable_id("transaction", source_row)
    transactions.append({
        "id": transaction_id,
        "sourceRow": source_row,
        "date": transaction_date.date().isoformat(),
        "type": transaction_type,
        "category": major,
        "minorCategory": minor or None,
        "memo": memo,
        "amount": int(round(float(amount))),
        "accountName": primary_name,
        "toAccountName": secondary_name or None,
        "isFixed": bool(categories.get(category_key, False)),
        "transferGroupId": stable_id("transfer", source_row) if transaction_type == "transfer" else None,
    })

for name in sorted(missing_account_names):
    if name in account_by_name:
        continue
    accounts.append({
        "id": stable_id("account", name),
        "accountCode": f"import-{len(accounts) + 1:03d}",
        "name": name,
        "type": "bank",
        "classification": "asset",
        "majorCategory": "💰현금·예금",
        "minorCategory": "보통예금",
        "openingBalance": 0,
        "memo": "원본 가계부에서 계좌가 비어 있거나 설정에 없는 거래를 보존하기 위해 생성됨",
        "isHidden": False,
        "paymentDay": None,
        "color": "#777b86",
    })

account_by_name = {account["name"]: account for account in accounts}
for transaction in transactions:
    transaction["accountId"] = account_by_name[transaction.pop("accountName")]["id"]
    destination = transaction.pop("toAccountName")
    transaction["toAccountId"] = account_by_name[destination]["id"] if destination else None

category_rows = [
    {
        "id": stable_id("category", f"{kind}|{major}|{minor}"),
        "transactionType": kind,
        "majorCategory": major,
        "minorCategory": minor,
        "isFixed": fixed,
        "sortOrder": index,
        "isHidden": False,
    }
    for index, ((kind, major, minor), fixed) in enumerate(categories.items(), start=1)
]

payload = {
    "source": SOURCE.name,
    "settings": {"startYear": start_year, "fiscalStartMonth": fiscal_start_month},
    "accounts": accounts,
    "categories": category_rows,
    "transactions": transactions,
    "summary": {
        "accountCount": len(accounts),
        "categoryCount": len(category_rows),
        "transactionCount": len(transactions),
        "income": sum(item["amount"] for item in transactions if item["type"] == "income"),
        "expense": sum(item["amount"] for item in transactions if item["type"] == "expense"),
        "transfer": sum(item["amount"] for item in transactions if item["type"] == "transfer"),
        "missingAccountNames": sorted(missing_account_names),
    },
}

OUTPUT.parent.mkdir(exist_ok=True)
OUTPUT.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
print(json.dumps(payload["summary"], ensure_ascii=False, indent=2))
